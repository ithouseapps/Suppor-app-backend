from django.utils import timezone
from django.utils.timezone import localtime
from django.db.models import Count, Q, F
from datetime import datetime
from rest_framework import viewsets, status, generics
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    User, Room, Subject, AcademicSupport,
    SupportSchedule, Booking, Lesson, LessonRating, SupportBan, BotConfig
)
from .serializers import (
    LoginSerializer, UserSerializer, RoomSerializer,
    SubjectSerializer, AcademicSupportSerializer,
    AcademicSupportSimpleSerializer, SupportScheduleSerializer,
    BookingSerializer, LessonSerializer, LessonRatingSerializer,
    StartLessonSerializer, EndLessonSerializer,
    SupportDashboardSerializer, SupportBanSerializer
)
from .permissions import IsAdmin, IsSupport, IsStudent, IsAdminOrSupport
from .bot import notify_support_free, notify_support_busy


@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def setup_view(request):
    if User.objects.filter(is_superuser=True).exists():
        return Response({'message': 'Admin already exists'})
    User.objects.create_superuser(username='admin', password='admin123', role='admin')
    return Response({'message': 'Admin created: admin / admin123'})

@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = serializer.validated_data
    refresh = RefreshToken.for_user(user)
    return Response({
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'user': UserSerializer(user).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    user = request.user
    data = UserSerializer(user).data
    if user.role == 'support':
        try:
            support = AcademicSupport.objects.get(user=user)
            data['support_profile'] = AcademicSupportSerializer(support).data
        except AcademicSupport.DoesNotExist:
            pass
    return Response(data)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):
        if self.action == 'me':
            self.permission_classes = [IsAuthenticated]
        elif self.request.method == 'POST':
            role = self.request.data.get('role', 'student')
            if role in ('admin', 'support'):
                self.permission_classes = [IsAdmin]
            else:
                self.permission_classes = [AllowAny]
        elif self.request.method in ['GET', 'HEAD', 'OPTIONS']:
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    def get_queryset(self):
        qs = User.objects.all()
        role = self.request.query_params.get('role')
        if role:
            qs = qs.filter(role=role)
        return qs

    @action(detail=False, methods=['GET', 'PATCH'])
    def me(self, request):
        if request.method == 'GET':
            return Response(UserSerializer(request.user).data)
        serializer = UserSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()


class AcademicSupportViewSet(viewsets.ModelViewSet):
    queryset = AcademicSupport.objects.all()
    serializer_class = AcademicSupportSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()


class SupportScheduleViewSet(viewsets.ModelViewSet):
    queryset = SupportSchedule.objects.all()
    serializer_class = SupportScheduleSerializer
    permission_classes = [IsAdminOrSupport]

    def get_queryset(self):
        qs = SupportSchedule.objects.all()
        user = self.request.user
        if user.role == 'support':
            try:
                support = AcademicSupport.objects.get(user=user)
                qs = qs.filter(support=support)
            except AcademicSupport.DoesNotExist:
                qs = qs.none()
        return qs

    def perform_create(self, serializer):
        if self.request.user.role == 'support':
            support = AcademicSupport.objects.get(user=self.request.user)
            serializer.save(support=support)
        elif 'support' in self.request.data:
            serializer.save()
        else:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'support': 'This field is required.'})


class BookingViewSet(viewsets.ModelViewSet):
    queryset = Booking.objects.all()
    serializer_class = BookingSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            self.permission_classes = [IsStudent]
        elif self.request.method == 'GET':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        qs = Booking.objects.all()
        if user.role == 'student':
            qs = qs.filter(student=user)
        elif user.role == 'support':
            try:
                support = AcademicSupport.objects.get(user=user)
                qs = qs.filter(support=support)
            except AcademicSupport.DoesNotExist:
                qs = qs.none()
        date = self.request.query_params.get('date')
        support_id = self.request.query_params.get('support')
        status = self.request.query_params.get('status')
        if date:
            qs = qs.filter(date=date)
        if support_id:
            qs = qs.filter(support_id=support_id)
        if status:
            qs = qs.filter(status=status)
        return qs.order_by('date', 'start_time')


class LessonViewSet(viewsets.ModelViewSet):
    queryset = Lesson.objects.all()
    serializer_class = LessonSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            self.permission_classes = [IsAdminOrSupport]
        elif self.request.method == 'GET':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        qs = Lesson.objects.all()
        if user.role == 'student':
            qs = qs.filter(student=user)
        elif user.role == 'support':
            try:
                support = AcademicSupport.objects.get(user=user)
                qs = qs.filter(support=support)
            except AcademicSupport.DoesNotExist:
                qs = qs.none()
        date = self.request.query_params.get('date')
        support_id = self.request.query_params.get('support')
        if date:
            qs = qs.filter(start_time__date=date)
        if support_id:
            qs = qs.filter(support_id=support_id)
        return qs.order_by('-start_time')


@api_view(['POST'])
@permission_classes([IsSupport])
def start_lesson(request):
    serializer = StartLessonSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    try:
        support = AcademicSupport.objects.get(user=request.user)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support profili topilmadi'}, status=400)

    if support.is_banned:
        return Response({'error': 'Siz bloklangan support siz. Admin bilan boglaning.'}, status=400)

    # Auto-end stale active lessons (older than 12 hours)
    from datetime import timedelta
    stale_cutoff = timezone.now() - timedelta(hours=12)
    Lesson.objects.filter(support=support, is_active=True, start_time__lt=stale_cutoff).update(
        is_active=False, end_time=timezone.now()
    )

    # Check if already in a lesson
    active = Lesson.objects.filter(support=support, is_active=True).first()
    if active:
        return Response({'error': 'Siz hozir dars o\'tyapsiz. Avval darsni tugating.'}, status=400)

    student_name = serializer.validated_data['student_name'].strip()
    student = User.objects.filter(username__iexact=student_name, role='student').first()
    if not student:
        student = User.objects.create(
            username=student_name.lower().replace(' ', '_'),
            first_name=student_name,
            role='student'
        )

    subject = None
    subject_id = serializer.validated_data.get('subject_id')
    if subject_id:
        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            pass

    # Parse scheduled_start if provided
    scheduled_start = None
    scheduled_start_str = serializer.validated_data.get('scheduled_start')
    if scheduled_start_str:
        try:
            now = timezone.now()
            parsed = datetime.strptime(scheduled_start_str, '%H:%M')
            scheduled_start = now.replace(hour=parsed.hour, minute=parsed.minute, second=0, microsecond=0)
        except ValueError:
            pass

    now = timezone.now()

    # Find room from active schedule
    current_time = now.time()
    day = now.weekday()
    schedule = SupportSchedule.objects.filter(
        support=support,
        day_of_week=day,
        start_time__lte=current_time,
        end_time__gte=current_time
    ).first()

    lesson = Lesson.objects.create(
        support=support,
        student=student,
        subject=subject,
        room=schedule.room if schedule else None,
        topic=serializer.validated_data['topic'],
        comment=serializer.validated_data.get('comment', ''),
        start_time=now,
        scheduled_start=scheduled_start,
        is_active=True,
        booking_id=serializer.validated_data.get('booking_id'),
        student_count=serializer.validated_data.get('student_count')
    )

    # Update booking if exists
    booking_id = serializer.validated_data.get('booking_id')
    if booking_id:
        Booking.objects.filter(id=booking_id).update(status='completed')

    try:
        notify_support_busy(
            support_name=lesson.support.user.username,
            student_name=lesson.student.username,
            topic=lesson.topic,
        )
    except Exception:
        pass

    return Response(LessonSerializer(lesson).data, status=201)


@api_view(['POST'])
@permission_classes([IsSupport])
def end_lesson(request):
    serializer = EndLessonSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    try:
        lesson = Lesson.objects.get(
            id=serializer.validated_data['lesson_id'],
            is_active=True
        )
    except Lesson.DoesNotExist:
        return Response({'error': 'Dars topilmadi yoki allaqachon tugatilgan'}, status=400)

    # Ensure the support owns this lesson
    if lesson.support.user != request.user:
        return Response({'error': 'Bu dars sizniki emas'}, status=403)

    lesson.end_time = timezone.now()
    lesson.is_active = False
    lesson.save()

    try:
        notify_support_free(
            support_name=lesson.support.user.username,
            student_name=lesson.student.username,
            topic=lesson.topic,
        )
    except Exception:
        pass

    return Response(LessonSerializer(lesson).data)


@api_view(['POST'])
@permission_classes([IsStudent])
def create_booking(request):
    support_id = request.data.get('support')
    subject_id = request.data.get('subject')
    date = request.data.get('date')
    start_time = request.data.get('start_time')
    end_time = request.data.get('end_time')

    if not all([support_id, date, start_time, end_time]):
        return Response({'error': 'support, date, start_time, end_time kerak'}, status=400)

    try:
        support = AcademicSupport.objects.get(id=support_id)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support topilmadi'}, status=400)

    from datetime import datetime as dt, timedelta as td, time as t

    try:
        date = dt.strptime(date, '%Y-%m-%d').date()
        start_time_obj = dt.strptime(start_time, '%H:%M').time()
        end_time_obj = dt.strptime(end_time, '%H:%M').time()
    except ValueError:
        return Response({'error': 'Sana yoki vaqt formati noto\'g\'ri'}, status=400)

    # Check for scheduling conflicts (bookings)
    existing_booking = Booking.objects.filter(
        support=support,
        date=date,
        start_time__lt=end_time_obj,
        end_time__gt=start_time_obj,
        status__in=['pending', 'confirmed']
    ).exists()
    if existing_booking:
        return Response({'error': 'Bu vaqt band'}, status=400)

    # Check for active lessons at this time
    requested_start = timezone.make_aware(dt.strptime(f"{date} {start_time}", '%Y-%m-%d %H:%M'))
    requested_end = timezone.make_aware(dt.strptime(f"{date} {end_time}", '%Y-%m-%d %H:%M'))
    active_lessons = Lesson.objects.filter(
        support=support,
        is_active=True
    )
    for lesson in active_lessons:
        less_start = lesson.start_time
        less_end = lesson.end_time or lesson.start_time + td(hours=1)
        if less_start < requested_end and less_end > requested_start:
            return Response({'error': 'Bu vaqt band (dars davom etmoqda)'}, status=400)

    booking = Booking.objects.create(
        support=support,
        student=request.user,
        subject_id=subject_id,
        date=date,
        start_time=start_time_obj,
        end_time=end_time_obj,
        status='pending'
    )

    return Response(BookingSerializer(booking).data, status=201)


@api_view(['POST'])
@permission_classes([IsSupport])
def cancel_booking(request, booking_id):
    try:
        booking = Booking.objects.get(id=booking_id)
    except Booking.DoesNotExist:
        return Response({'error': 'Booking topilmadi'}, status=404)

    if booking.support.user != request.user:
        return Response({'error': 'Bu booking sizniki emas'}, status=403)

    booking.status = 'cancelled'
    booking.save()
    return Response(BookingSerializer(booking).data)


@api_view(['POST'])
@permission_classes([IsSupport])
def confirm_booking(request, booking_id):
    try:
        booking = Booking.objects.get(id=booking_id)
    except Booking.DoesNotExist:
        return Response({'error': 'Booking topilmadi'}, status=404)

    if booking.support.user != request.user:
        return Response({'error': 'Bu booking sizniki emas'}, status=403)

    booking.status = 'confirmed'
    booking.save()
    return Response(BookingSerializer(booking).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    today = timezone.now().date()

    today_lessons = Lesson.objects.filter(start_time__date=today)
    today_bookings = Booking.objects.filter(date=today)

    total_students = today_lessons.filter(
        end_time__isnull=False
    ).values('student').distinct().count()

    total_lessons = today_lessons.filter(end_time__isnull=False).count()
    total_bookings = today_bookings.count()

    active_lessons = Lesson.objects.filter(is_active=True)
    busy_ids = active_lessons.values_list('support_id', flat=True)
    all_supports = AcademicSupport.objects.filter(is_active=True)
    free = all_supports.exclude(id__in=busy_ids).count()
    busy = all_supports.filter(id__in=busy_ids).count()

    return Response({
        'today_students': total_students,
        'today_lessons': total_lessons,
        'today_bookings': total_bookings,
        'free_supports': free,
        'busy_supports': busy,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_supports(request):
    supports = AcademicSupport.objects.filter(is_active=True).prefetch_related(
        'user', 'subjects', 'lessons'
    )
    serializer = SupportDashboardSerializer(supports, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_rooms(request):
    rooms = Room.objects.filter(is_active=True)
    now = timezone.now()
    current_time = now.time()
    day = now.weekday()

    data = []
    for room in rooms:
        active_lesson = Lesson.objects.filter(
            room=room,
            is_active=True
        ).select_related('support__user', 'student', 'subject').first()

        room_data = {
            'id': room.id,
            'name': room.name,
            'status': 'busy' if active_lesson else 'free',
        }

        if active_lesson:
            room_data['current_lesson'] = {
                'support_name': active_lesson.support.user.username,
                'student_name': active_lesson.student.username,
                'subject': active_lesson.subject.name if active_lesson.subject else None,
                'topic': active_lesson.topic,
                'start_time': localtime(active_lesson.start_time).strftime('%H:%M'),
            }

        data.append(room_data)

    return Response(data)


@api_view(['GET'])
@permission_classes([IsStudent])
def available_slots(request, support_id):
    date_str = request.query_params.get('date')
    if not date_str:
        return Response({'error': 'date kerak (YYYY-MM-DD)'}, status=400)

    from datetime import datetime, timedelta
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    day = date.weekday()

    try:
        support = AcademicSupport.objects.get(id=support_id)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support topilmadi'}, status=400)

    schedules = SupportSchedule.objects.filter(support=support, day_of_week=day)

    # Get existing bookings for the day (pending/confirmed only)
    bookings = Booking.objects.filter(
        support=support,
        date=date,
        status__in=['pending', 'confirmed']
    )

    # Get active lessons for this support (check across UTC date boundary)
    day_start = timezone.make_aware(datetime.combine(date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)
    active_lessons = Lesson.objects.filter(
        support=support,
        is_active=True,
        start_time__lt=day_end,
    )

    slots = []
    for schedule in schedules:
        current = datetime.combine(date, schedule.start_time)
        end = datetime.combine(date, schedule.end_time)

        while current + timedelta(minutes=30) <= end:
            slot_start = current.time()
            slot_end = (current + timedelta(minutes=30)).time()

            is_booked = bookings.filter(
                start_time__lt=slot_end,
                end_time__gt=slot_start
            ).exists()

            # Also check if there's an active lesson at this time
            if not is_booked:
                slot_start_dt = timezone.make_aware(datetime.combine(date, slot_start))
                slot_end_dt = timezone.make_aware(datetime.combine(date, slot_end))
                for lesson in active_lessons:
                    less_end = lesson.end_time or lesson.start_time + timedelta(hours=1)
                    if lesson.start_time < slot_end_dt and less_end > slot_start_dt:
                        is_booked = True
                        break

            slots.append({
                'start_time': slot_start.strftime('%H:%M'),
                'end_time': slot_end.strftime('%H:%M'),
                'is_booked': is_booked,
                'room': schedule.room.name if schedule.room else None,
            })

            current += timedelta(minutes=30)

    return Response(slots)


@api_view(['GET'])
@permission_classes([IsSupport])
def support_profile_view(request):
    try:
        support = AcademicSupport.objects.get(user=request.user)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support profile topilmadi'}, status=400)

    # All lessons by this support, grouped by student
    lessons = Lesson.objects.filter(support=support).select_related(
        'student', 'subject', 'room'
    ).order_by('-start_time')

    student_lessons = {}
    total_minutes = 0
    total_completed = 0
    for lesson in lessons:
        sid = lesson.student.id
        if sid not in student_lessons:
            student_lessons[sid] = {
                'student': {
                    'id': lesson.student.id,
                    'username': lesson.student.username,
                    'first_name': lesson.student.first_name,
                    'last_name': lesson.student.last_name,
                },
                'total_lessons': 0,
                'completed_lessons': 0,
                'total_minutes': 0,
                'lessons': [],
            }
        student_lessons[sid]['total_lessons'] += 1
        if not lesson.is_active:
            student_lessons[sid]['completed_lessons'] += 1
            total_completed += 1
            if lesson.end_time:
                delta = lesson.end_time - lesson.start_time
                mins = int(delta.total_seconds() / 60)
                student_lessons[sid]['total_minutes'] += mins
                total_minutes += mins
        student_lessons[sid]['lessons'].append(LessonSerializer(lesson).data)

    return Response({
        'support': AcademicSupportSerializer(support).data,
        'students': list(student_lessons.values()),
        'total_students': len(student_lessons),
        'total_lessons': lessons.count(),
        'total_completed': total_completed,
        'total_minutes': total_minutes,
    })


@api_view(['GET'])
@permission_classes([IsStudent])
def student_profile_view(request):
    lessons = Lesson.objects.filter(student=request.user).select_related(
        'support__user', 'subject', 'room'
    ).order_by('-start_time')

    total_completed = 0
    total_minutes = 0
    for lesson in lessons:
        if not lesson.is_active:
            total_completed += 1
            if lesson.end_time:
                delta = lesson.end_time - lesson.start_time
                total_minutes += int(delta.total_seconds() / 60)

    return Response({
        'user': UserSerializer(request.user).data,
        'total_lessons': lessons.count(),
        'total_completed': total_completed,
        'total_minutes': total_minutes,
    })


@api_view(['GET'])
@permission_classes([IsStudent])
def support_list_for_student(request):
    supports = AcademicSupport.objects.filter(is_active=True).prefetch_related('user', 'subjects')
    serializer = AcademicSupportSimpleSerializer(supports, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsSupport])
def student_status_list(request):
    today = timezone.now().date()
    active_lessons = Lesson.objects.filter(
        is_active=True
    ).select_related('student', 'support__user')

    busy_student_ids = set(active_lessons.values_list('student_id', flat=True))

    students = User.objects.filter(role='student')
    data = []
    for student in students:
        is_busy = student.id in busy_student_ids
        current_lesson = None
        if is_busy:
            lesson = active_lessons.filter(student=student).first()
            current_lesson = {
                'support_name': lesson.support.user.username if lesson else None,
                'topic': lesson.topic if lesson else None,
                'start_time': localtime(lesson.start_time).strftime('%H:%M') if lesson else None,
            }
        data.append({
            'id': student.id,
            'username': student.username,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'email': student.email,
            'phone': student.phone,
            'secret_id': student.secret_id,
            'is_busy': is_busy,
            'current_lesson': current_lesson,
        })
    return Response(data)


@api_view(['POST'])
@permission_classes([IsStudent])
def rate_lesson(request):
    lesson_id = request.data.get('lesson_id')
    score = request.data.get('score')
    comment = request.data.get('comment', '')

    if not lesson_id or not score:
        return Response({'error': 'lesson_id va score kerak'}, status=400)

    try:
        lesson = Lesson.objects.get(id=lesson_id, is_active=False)
    except Lesson.DoesNotExist:
        return Response({'error': 'Dars topilmadi'}, status=400)

    if lesson.student != request.user:
        return Response({'error': 'Bu dars sizniki emas'}, status=403)

    if LessonRating.objects.filter(lesson=lesson).exists():
        return Response({'error': 'Siz bu darsni baholagansiz'}, status=400)

    rating = LessonRating.objects.create(
        lesson=lesson,
        student=request.user,
        score=score,
        comment=comment
    )

    return Response(LessonRatingSerializer(rating).data, status=201)


@api_view(['POST'])
@permission_classes([IsAdmin])
def ban_support(request, support_id):
    reason = request.data.get('reason', '')

    try:
        support = AcademicSupport.objects.get(id=support_id)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support topilmadi'}, status=400)

    # End any active lessons
    Lesson.objects.filter(support=support, is_active=True).update(is_active=False, end_time=timezone.now())

    SupportBan.objects.create(
        support=support,
        banned_by=request.user,
        reason=reason,
        is_active=True
    )
    support.is_banned = True
    support.save()

    return Response({'message': 'Support bloklandi'})


@api_view(['POST'])
@permission_classes([IsAdmin])
def unban_support(request, support_id):
    try:
        support = AcademicSupport.objects.get(id=support_id)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support topilmadi'}, status=400)

    active_ban = SupportBan.objects.filter(support=support, is_active=True).first()
    if active_ban:
        active_ban.is_active = False
        active_ban.lifted_at = timezone.now()
        active_ban.save()

    support.is_banned = False
    support.save()

    return Response({'message': 'Support blokdan ochildi'})


@api_view(['GET'])
@permission_classes([IsAdmin])
def support_delays(request):
    supports = AcademicSupport.objects.filter(is_active=True).prefetch_related('user', 'lessons')

    data = []
    for sup in supports:
        lessons = Lesson.objects.filter(support=sup, is_active=False, scheduled_start__isnull=False)
        total_delays = lessons.filter(start_time__gt=F('scheduled_start')).count()
        total_early = lessons.filter(start_time__lt=F('scheduled_start')).count()
        total_on_time = lessons.filter(start_time=F('scheduled_start')).count()

        avg_delay = 0
        delayed_lessons = lessons.filter(start_time__gt=F('scheduled_start'))
        if delayed_lessons.exists():
            total_delay_minutes = 0
            count = 0
            for l in delayed_lessons:
                delta = l.start_time - l.scheduled_start
                total_delay_minutes += int(delta.total_seconds() / 60)
                count += 1
            avg_delay = total_delay_minutes / count if count else 0

        data.append({
            'id': sup.id,
            'username': sup.user.username,
            'total_lessons': lessons.count(),
            'delayed_lessons': total_delays,
            'early_lessons': total_early,
            'on_time_lessons': total_on_time,
            'avg_delay_minutes': round(avg_delay, 1),
            'is_banned': sup.is_banned,
        })

    return Response(data)


@api_view(['GET'])
@permission_classes([IsStudent])
def support_schedules_for_student(request):
    supports = AcademicSupport.objects.filter(is_active=True, is_banned=False).prefetch_related(
        'user', 'subjects', 'schedules__room'
    )
    result = []
    for sup in supports:
        schedule_data = []
        for sch in sup.schedules.all():
            schedule_data.append({
                'id': sch.id,
                'day_of_week': sch.day_of_week,
                'day_of_week_display': sch.get_day_of_week_display(),
                'start_time': sch.start_time.strftime('%H:%M'),
                'end_time': sch.end_time.strftime('%H:%M'),
                'room_name': sch.room.name if sch.room else None,
            })
        result.append({
            'id': sup.id,
            'user': {
                'id': sup.user.id,
                'username': sup.user.username,
                'first_name': sup.user.first_name,
                'last_name': sup.user.last_name,
            },
            'subjects': [{'id': s.id, 'name': s.name} for s in sup.subjects.all()],
            'schedules': schedule_data,
        })
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_support_profile_view(request, support_id):
    try:
        support = AcademicSupport.objects.get(id=support_id)
    except AcademicSupport.DoesNotExist:
        return Response({'error': 'Support topilmadi'}, status=400)

    lessons = Lesson.objects.filter(support=support).select_related(
        'student', 'subject', 'room'
    ).order_by('-start_time')

    student_lessons = {}
    total_minutes = 0
    total_completed = 0
    for lesson in lessons:
        sid = lesson.student.id
        if sid not in student_lessons:
            student_lessons[sid] = {
                'student': {
                    'id': lesson.student.id,
                    'username': lesson.student.username,
                    'first_name': lesson.student.first_name,
                    'last_name': lesson.student.last_name,
                },
                'total_lessons': 0,
                'completed_lessons': 0,
                'total_minutes': 0,
                'lessons': [],
            }
        student_lessons[sid]['total_lessons'] += 1
        if not lesson.is_active:
            student_lessons[sid]['completed_lessons'] += 1
            total_completed += 1
            if lesson.end_time:
                delta = lesson.end_time - lesson.start_time
                mins = int(delta.total_seconds() / 60)
                student_lessons[sid]['total_minutes'] += mins
                total_minutes += mins
        student_lessons[sid]['lessons'].append(LessonSerializer(lesson).data)

    return Response({
        'support': AcademicSupportSerializer(support).data,
        'students': list(student_lessons.values()),
        'total_students': len(student_lessons),
        'total_lessons': lessons.count(),
        'total_completed': total_completed,
        'total_minutes': total_minutes,
    })


@api_view(['POST'])
@permission_classes([IsAdmin])
def generate_student_secret_id(request, student_id):
    import random
    import string
    try:
        student = User.objects.get(id=student_id, role='student')
    except User.DoesNotExist:
        return Response({'error': 'Student topilmadi'}, status=400)

    secret = ''.join(random.choices(string.digits, k=6))
    student.secret_id = secret
    student.save()

    return Response({'secret_id': secret, 'student_id': student.id})


from django.http import HttpResponse


@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_monthly_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    today = timezone.now()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))

    month_start = timezone.make_aware(datetime(year, month, 1))
    month_end = timezone.make_aware(datetime(year + 1, 1, 1)) if month == 12 else timezone.make_aware(datetime(year, month + 1, 1))

    supports = AcademicSupport.objects.filter(is_active=True).order_by('user__username')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Hisobot {month}.{year}"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_fill2 = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    support_data = []
    for support in supports:
        lessons = Lesson.objects.filter(
            support=support,
            start_time__gte=month_start,
            start_time__lt=month_end,
            is_active=False
        )
        total_lessons = lessons.count()
        student_ids = lessons.values_list('student_id', flat=True).distinct()
        total_students = student_ids.count()
        student_users = User.objects.filter(id__in=list(student_ids), role='student') if student_ids else []
        active_students = ', '.join(s.get_full_name() or s.username for s in student_users)
        total_group_count = sum(l.student_count or 0 for l in lessons)

        support_data.append({
            'name': support.user.get_full_name() or support.user.username,
            'total_lessons': total_lessons,
            'total_students': total_students,
            'active_students': active_students,
            'total_group_count': total_group_count,
        })

    cell = ws.cell(row=1, column=1, value=f"{month}.{year} hisobot")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    for i, sd in enumerate(support_data, 2):
        cell = ws.cell(row=1, column=i, value=sd['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    row_labels = ['Jami darslar', 'Jami studentlar', 'Guruhli studentlar', 'Faol studentlar']
    fills = [header_fill, header_fill2, header_fill2, header_fill]
    for r, (label, fill) in enumerate(zip(row_labels, fills), 2):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        for i, sd in enumerate(support_data, 2):
            if r == 2:
                val = sd['total_lessons']
            elif r == 3:
                val = sd['total_students']
            elif r == 4:
                val = sd['total_group_count']
            else:
                val = sd['active_students']
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    from openpyxl.utils import get_column_letter
    ws.column_dimensions['A'].width = 20
    for i in range(2, len(support_data) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monthly_report_{year}_{month:02d}.xlsx"'
    wb.save(response)
    return response


@api_view(['GET', 'PUT'])
@permission_classes([IsAdmin])
def bot_config_view(request):
    config = BotConfig.objects.first()
    if not config:
        config = BotConfig.objects.create()

    if request.method == 'PUT':
        chat_id = request.data.get('chat_id')
        if chat_id:
            config.chat_id = str(chat_id)

        bot_token = request.data.get('bot_token')
        if bot_token:
            config.bot_token = str(bot_token)

        config.save()
        return Response({'chat_id': config.chat_id, 'bot_token': config.bot_token, 'message': 'Saqlandi'})

    return Response({'chat_id': config.chat_id, 'bot_token': config.bot_token})
